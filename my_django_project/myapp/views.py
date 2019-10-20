from django.shortcuts import render
import re
import xlrd
import MySQLdb
import openpyxl

from datetime import date
import datetime

from .models import EmployeeInfoData

def emailvalid(emailid):
    # Make a regular expression for validating an Email
    regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
    return re.search(regex,emailid)

def isphoneValid(number): 
      
    # 1) Begins with 0 or 91 
    # 2) Then contains 7 or 8 or 9. 
    # 3) Then contains 9 digits 
    Pattern = re.compile("(0/91)?[7-9][0-9]{9}") 
    return Pattern.match(number) 

def calculate_age(born):
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))

def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        if "upload" in request.POST:
            excel_file = request.FILES["excel_file"]
            book = xlrd.open_workbook(file_contents=excel_file.read())
            
            sheet = book.sheet_by_index(0)
            # Establish a MySQL connection
            database = MySQLdb.connect (host="localhost", user = "root", passwd = "Basavaraj", db = "excel_data")
            cursor = database.cursor()
            query = """INSERT INTO myapp_employeeinfodata (name, email, phonenumber, age) VALUES (%s, %s, %s, %s)"""

            # Create a For loop to iterate through each row in the XLS file
            for r in range(1, sheet.nrows):
                name     = sheet.cell(r,0).value
     
                if sheet.cell(r,1).value:
                    email_valid = emailvalid(sheet.cell(r,1).value)
                    if emailvalid:
                        email = sheet.cell(r,1).value
                    else:
                        email = ''
                if sheet.cell(r,2).value:
                    phone_valid = isphoneValid(str(sheet.cell(r,2).value))      
                    if phone_valid:
                        phonenumber = sheet.cell(r,2).value
                    else:
                        phonenumber = ''

                if sheet.cell(r,3).value:
                    age = calculate_age(datetime.datetime(*xlrd.xldate_as_tuple(sheet.cell(r,3).value, book.datemode)))
                try:
                    values = (name, email, phonenumber, age)
                    cursor.execute(query, values)
                except Exception as e:
                    print("")
            cursor.close()
            # Commit the transaction
            database.commit()

            # Close the database connection
            database.close()
             # Execute sql Query
            excel_data = EmployeeInfoData.objects.all()
        elif 'search' in request.POST:
            search_name = request.POST.get('search_value')
            excel_data = EmployeeInfoData.objects.filter(name=search_name)
        else:
            excel_data = EmployeeInfoData.objects.all()

        return render(request, 'myapp/index.html', {"excel_data":excel_data})









